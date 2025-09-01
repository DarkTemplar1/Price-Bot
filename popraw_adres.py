# -*- coding: utf-8 -*-
import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import tkinter as tk
from tkinter import messagebox

from openpyxl import load_workbook, Workbook

APP_TITLE = "popraw_adres – automatyczne uzupełnianie adresów"

TERYT_DEFAULT = "TERYT.xlsx"
NR_KW_DEFAULT = "Nr KW.xlsx"

# ========================= pomocnicze =========================

DIAC_MAP = str.maketrans({
    "ą":"a","ć":"c","ę":"e","ł":"l","ń":"n","ó":"o","ś":"s","ź":"z","ż":"z",
    "Ą":"A","Ć":"C","Ę":"E","Ł":"L","Ń":"N","Ó":"O","Ś":"S","Ź":"Z","Ż":"Z",
})

def norm_key(s: str) -> str:
    if s is None: return ""
    s = str(s).strip().lower()
    s = re.sub(r"^(wojew[oó]dztwo|woj\.)\s+", "", s)  # usuń "województwo " / "woj. "
    return s.translate(DIAC_MAP)

def _info(msg: str):
    try:
        r = tk.Tk(); r.withdraw()
        messagebox.showinfo("Informacja", msg, parent=r)
        r.destroy()
    except Exception:
        print(msg)

def _err(msg: str):
    try:
        r = tk.Tk(); r.withdraw()
        messagebox.showerror("Błąd", msg, parent=r)
        r.destroy()
    except Exception:
        print("BŁĄD:", msg, file=sys.stderr)

# ========================= TERYT ==============================

def load_teryt(path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["TERYT"] if "TERYT" in wb.sheetnames else wb.active
    headers = {str(c.value).strip(): i for i, c in enumerate(ws[1], start=1) if c.value}
    req = ["Województwo","Powiat","Gmina","Miejscowość","Dzielnica"]
    for r in req:
        if r not in headers:
            raise RuntimeError(f"Brak kolumny '{r}' w {path}")
    rows = []
    for r in range(2, ws.max_row+1):
        rows.append({k: (ws.cell(row=r, column=headers[k]).value or "") for k in req})
    return rows

def filter_candidates(teryt_rows, w=None, p=None, g=None, m=None, d=None):
    w = norm_key(w) if w else ""
    p = norm_key(p) if p else ""
    g = norm_key(g) if g else ""
    m = norm_key(m) if m else ""
    d = norm_key(d) if d else ""
    out = []
    for rec in teryt_rows:
        ok = True
        if w and norm_key(rec["Województwo"]) != w: ok = False
        if p and norm_key(rec["Powiat"]) != p: ok = False
        if g and norm_key(rec["Gmina"]) != g: ok = False
        if m and norm_key(rec["Miejscowość"]) != m: ok = False
        if d and norm_key(rec["Dzielnica"]) != d: ok = False
        if ok: out.append(rec)
    return out

def stable_values(cands):
    fields = ["Województwo","Powiat","Gmina","Miejscowość","Dzielnica"]
    out = {}
    for f in fields:
        vals = {rec[f] for rec in cands if str(rec[f]).strip() != ""}
        if len(vals) == 1:
            out[f] = next(iter(vals))
    return out

# ===================== Nr KW.xlsx (prefiks→woj.) =====================

def _strip_accents_lower(s: str) -> str:
    return (s or "").translate(DIAC_MAP).lower().strip()

def load_kw_voiv_mapping(path: Path) -> Dict[str, str]:
    """
    Odczytuje mapę prefiksu (pierwsze 4 znaki z Nr KW) → Województwo.
    Obsługa elastyczna nagłówków:
    - prefiks: jeden z ['prefiks','prefix','kod','kod sadu','kod sądu','oznaczenie','nr kw','nr_kw','wydzial']
    - województwo: nagłówek zawierający 'woj' (po usunięciu polskich znaków)
    Jeśli nie znajdzie nagłówków, a są co najmniej 2 kolumny — używa kolumn (1→prefiks, 2→woj.).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws.max_row < 2:
        return {}
    # spróbuj zidentyfikować kolumny po nagłówkach
    pref_cand = {"prefiks","prefix","kod","kod sadu","kod sądu","oznaczenie","nr kw","nr_kw","wydzial","wydział"}
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    pref_col = None
    woj_col = None
    for i, h in enumerate(headers, start=1):
        hnorm = _strip_accents_lower(h)
        if pref_col is None and hnorm in pref_cand:
            pref_col = i
        if woj_col is None and "woj" in hnorm:
            woj_col = i
    if pref_col is None or woj_col is None:
        # fallback: pierwsze dwie kolumny
        pref_col, woj_col = 1, 2

    out = {}
    for r in range(2, ws.max_row+1):
        pref = ws.cell(row=r, column=pref_col).value
        woj = ws.cell(row=r, column=woj_col).value
        if pref is None or woj is None:
            continue
        key = str(pref).strip().upper()
        if not key:
            continue
        out[key] = str(woj).strip()
    return out

def kw_prefix(nr_kw: str) -> str:
    if not nr_kw:
        return ""
    # pierwsze 4 alfanumeryczne znaki
    cleaned = "".join(ch for ch in str(nr_kw).strip() if ch.isalnum())
    return cleaned[:4].upper()

# =================== obsługa Excela z raportem ===================

ADDR_COLS = ["Województwo","Powiat","Gmina","Miejscowość","Dzielnica"]
VALUE_COLS = [
    "Średnia cena za m² (z bazy)",
    "Średnia skorygowana cena za m² (z bazy)",
    "Statystyczna wartość nieruchomości",
]
KW_COL = "Nr KW"

def find_or_create_column(ws, header: str) -> int:
    """Zwróć indeks kolumny dla nagłówka; jeśli nie ma — dodaj na końcu."""
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    for idx, h in enumerate(headers, start=1):
        if h == header:
            return idx
    col = len(headers) + 1
    ws.cell(row=1, column=col, value=header)
    return col

def _read_headers(ws) -> Dict[str, int]:
    return { (c.value or ""): i for i, c in enumerate(ws[1], start=1) }

def get_sheet_candidates(wb) -> List:
    wanted = []
    for name in ("raport", "raport_odfiltrowane"):
        if name in wb.sheetnames:
            wanted.append(wb[name])
    if not wanted:
        wanted = [wb.active]
    return wanted

def complete_row_from_teryt(row_vals: Dict[str,str], teryt_rows: List[Dict[str,str]]) -> Dict[str,str]:
    """Próby dopasowania: pełne → (woj+miejsc) → (miejsc) → (gmina). Zwraca wartości do nadpisania."""
    w, p, g, m, d = [row_vals.get(k, "") for k in ADDR_COLS]
    # kolejno coraz luźniejsze filtry
    for args in [
        (w,p,g,m,d),
        (w,None,None,m,None),
        (None,None,None,m,None),
        (None,None,g,None,None),
    ]:
        cands = filter_candidates(teryt_rows, *args)
        if not cands:
            continue
        if len(cands) == 1:
            return cands[0]
        stab = stable_values(cands)
        if stab:
            return {**{k:"" for k in ADDR_COLS}, **stab}
    return {}

def process_workbook(in_xlsx: Path, teryt_xlsx: Path, nrkw_xlsx: Path) -> Tuple[int,int]:
    """
    Zwraca (ile_uzupełniono_adresem, ile_oznaczono_brak_adresu)
    """
    wb = load_workbook(in_xlsx)
    teryt_rows = load_teryt(teryt_xlsx)
    kw_map = load_kw_voiv_mapping(nrkw_xlsx) if nrkw_xlsx.exists() else {}

    updated_addr = 0
    marked_missing = 0

    for ws in get_sheet_candidates(wb):
        # upewnij się, że nagłówki istnieją
        hdr = _read_headers(ws)
        for col in [KW_COL] + ADDR_COLS + VALUE_COLS:
            if col not in hdr:
                find_or_create_column(ws, col)
        hdr = _read_headers(ws)

        col_idx = {k: hdr[k] for k in [KW_COL] + ADDR_COLS + VALUE_COLS}

        for r in range(2, ws.max_row + 1):
            cell = lambda name: ws.cell(row=r, column=col_idx[name])

            nr_kw_val = str(cell(KW_COL).value or "").strip()
            adr_vals = {k: str(cell(k).value or "").strip() for k in ADDR_COLS}

            # warunek "tylko Miejscowość"
            only_miejsc = (adr_vals["Miejscowość"] != "") and all(
                not adr_vals[k] for k in ["Województwo","Powiat","Gmina","Dzielnica"]
            )

            if only_miejsc and nr_kw_val and kw_map:
                pref = kw_prefix(nr_kw_val)
                woj = kw_map.get(pref, "")
                if woj:
                    ws.cell(row=r, column=col_idx["Województwo"], value=woj)
                    adr_vals["Województwo"] = woj  # do dopasowania TERYT

            # próba uzupełnienia z TERYT (jeśli cokolwiek mamy)
            filled = complete_row_from_teryt(adr_vals, teryt_rows)
            if filled:
                changed = False
                for k in ADDR_COLS:
                    newv = filled.get(k, "")
                    if newv and str(cell(k).value or "").strip() != str(newv).strip():
                        cell(k).value = newv
                        changed = True
                if changed:
                    updated_addr += 1

            # po wszystkich próbach: jeśli adres PUSTY wszędzie → wpisz "brak adresu" w 3 kolumnach
            adr_vals_after = {k: str(cell(k).value or "").strip() for k in ADDR_COLS}
            if all(v == "" for v in adr_vals_after.values()):
                for k in VALUE_COLS:
                    cell(k).value = "brak adresu"
                marked_missing += 1

    wb.save(in_xlsx)
    return updated_addr, marked_missing

# ============================= CLI =============================

def resolve_side_file(preferred_dir: Path, fallback_dir: Path, filename: str) -> Path:
    for base in (preferred_dir, fallback_dir):
        p = base / filename
        if p.exists():
            return p
    # zwróć ścieżkę w preferred_dir (może nie istnieć – przekażemy użytkownikowi komunikat)
    return preferred_dir / filename

def main():
    ap = argparse.ArgumentParser(description=APP_TITLE)
    ap.add_argument("--in", dest="infile", required=True, help="Plik Excel do poprawy (od Sortowni)")
    ap.add_argument("--teryt", dest="teryt", default=None, help="Ścieżka do TERYT.xlsx (opcjonalnie)")
    ap.add_argument("--nrkw", dest="nrkw", default=None, help="Ścieżka do 'Nr KW.xlsx' (opcjonalnie)")
    args = ap.parse_args()

    in_xlsx = Path(args.infile).expanduser().resolve()
    if not in_xlsx.exists():
        _err(f"Nie znaleziono pliku wejściowego: {in_xlsx}")
        sys.exit(2)

    # domyślne lokalizacje: najpierw folder pliku wejściowego, potem folder skryptu
    here = Path(__file__).resolve().parent
    teryt_path = Path(args.teryt).resolve() if args.teryt else resolve_side_file(in_xlsx.parent, here, TERYT_DEFAULT)
    nrkw_path  = Path(args.nrkw).resolve()  if args.nrkw  else resolve_side_file(in_xlsx.parent, here, NR_KW_DEFAULT)

    if not teryt_path.exists():
        _err(f"Brak pliku TERYT: {teryt_path}\nUmieść TERYT.xlsx obok pliku wejściowego lub wskaż --teryt.")
        sys.exit(3)

    if not nrkw_path.exists():
        _info(f"Uwaga: nie znaleziono '{NR_KW_DEFAULT}' ({nrkw_path}).\n"
              f"Uzupełnianie województwa z prefiksu Nr KW będzie pominięte.\n"
              f"Adres spróbujemy uzupełnić tylko na podstawie TERYT i miejscowości.")

    try:
        updated, missing = process_workbook(in_xlsx, teryt_path, nrkw_path)
    except Exception as e:
        _err(f"Nie udało się przetworzyć pliku:\n{e}")
        sys.exit(1)

    _info(f"Zakończono.\n"
          f"Uzupełniono adresy w wierszach: {updated}\n"
          f"Oznaczono 'brak adresu' w wierszach: {missing}\n\n"
          f"Plik: {in_xlsx}")

if __name__ == "__main__":
    main()
