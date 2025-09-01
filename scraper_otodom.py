#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import csv
import json
import re
import time
from http.client import RemoteDisconnected
from pathlib import Path

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook, load_workbook

# ------------------------------ WYJŚCIE: Desktop/Pulpit ------------------------------
def _detect_desktop() -> Path:
    home = Path.home()
    for name in ("Desktop", "Pulpit"):
        p = home / name
        if p.exists():
            return p
    return home

BASE_BAZA = _detect_desktop() / "baza danych"
BASE_LINKI = BASE_BAZA / "linki"
BASE_WOJ   = BASE_BAZA / "województwa"
BASE_LINKI.mkdir(parents=True, exist_ok=True)
BASE_WOJ.mkdir(parents=True, exist_ok=True)

# jeden wspólny plik dla wszystkich województw:
SINGLE_XLSX = BASE_WOJ / "wojewodztwa.xlsx"
# -------------------------------------------------------------------------------------

from adres_otodom import (
    set_contact_email,
    parsuj_adres_string,
    uzupelnij_braki_z_nominatim,
    dopelnij_powiat_gmina_jesli_brak,
    _clean_gmina,
    _consistency_pass_row,
)

CONTACT_EMAIL = "twoj_email@domena.pl"
WAIT_SECONDS = 20
set_contact_email(CONTACT_EMAIL)

szukane_pola = {
    "Cena": "cena",
    "Cena za m²": "cena_za_metr",
    "Powierzchnia": "metry",
    "Liczba pokoi": "liczba_pokoi",
    "Piętro": "pietro",
    "Rynek": "rynek",
    "Rok budowy": "rok_budowy",
    "Materiał budynku": "material",
}

LABEL_SYNONYMS = {
    "Liczba pokoi": ("liczba_pokoi", ["Liczba pokoi"]),
    "Rynek": ("rynek", ["Rynek", "Typ rynku"]),
    "Rok budowy": ("rok_budowy", ["Rok budowy", "Rok bud."]),
    "Materiał budynku": ("material", ["Materiał budynku", "Materiał", "Materiał bud."]),
    "Piętro": ("pietro", ["Piętro", "Kondygnacja", "Poziom"]),
}

KOLUMNY_WYJ = [
    "cena",
    "cena_za_metr",
    "metry",
    "liczba_pokoi",
    "pietro",
    "rynek",
    "rok_budowy",
    "material",
    "wojewodztwo",
    "powiat",
    "gmina",
    "miejscowosc",
    "dzielnica",
    "ulica",
    "link",
]

def format_pln_int(x) -> str | None:
    if x is None:
        return None
    try:
        return f"{int(x):,} zł".replace(",", " ")
    except Exception:
        return None

def fmt_metry(value) -> str | None:
    if value is None:
        return None
    try:
        v = float(str(value).replace(",", "."))
        s = f"{v:.2f}".rstrip("0").rstrip(".").replace(".", ",")
        return f"{s} m²"
    except Exception:
        return None

def _num(s: str):
    if not s:
        return None
    digits = re.sub(r"[^\d]", "", s)
    return int(digits) if digits else None

def oblicz_cena_za_metr(cena_str: str, metry_str: str) -> str:
    try:
        cena = int(re.sub(r"[^\d]", "", cena_str or ""))
        metry = float((metry_str or "").replace(",", ".").replace("m²", "").strip())
        if not cena or not metry:
            return ""
        wynik = round(cena / metry)
        return f"{wynik} zł/m²"
    except Exception:
        return ""

def _norm_txt(s: str | None) -> str:
    return re.sub(r"\s+", " ", (s or "").replace("\xa0", " ")).strip()

def _extract_by_labels_generic(soup, wynik: dict):
    labels = []
    label_to_key = {}
    for _canon, (key, syns) in LABEL_SYNONYMS.items():
        for s in syns:
            labels.append(re.escape(s))
            label_to_key[s.lower()] = key
    if not labels:
        return
    pat = re.compile(rf"^(?:{'|'.join(labels)})\s*:?\s*$", re.I)
    for node in soup.find_all(string=pat):
        lab_raw = _norm_txt(node)
        lab = re.sub(r":$", "", lab_raw, flags=re.I)
        key = label_to_key.get(lab.lower())
        if not key or wynik.get(key):
            continue
        val = None
        parent = node.parent
        for cand in (
            getattr(parent, "next_sibling", None),
            getattr(parent, "find_next_sibling", lambda: None)(),
            getattr(parent.parent if parent else None, "find_next_sibling", lambda: None)(),
        ):
            if hasattr(cand, "get_text"):
                t = _norm_txt(cand.get_text(" ", strip=True))
                if t and t.lower() != lab.lower():
                    val = t
                    break
        if not val:
            nxt = node.find_next(string=lambda t: _norm_txt(t) and _norm_txt(t).lower() != lab.lower())
            if nxt:
                val = _norm_txt(nxt)
        if val:
            wynik[key] = val

_REMOTE_CLOSED_MARKERS = (
    "remote end closed connection without response",
    "remotedisconnected",
    "connection reset",
    "chunked encoding",
)
def _is_remote_closed(exc: Exception) -> bool:
    s = (str(exc) or "").lower()
    if isinstance(exc, RemoteDisconnected):
        return True
    return any(m in s for m in _REMOTE_CLOSED_MARKERS)

def _safe_enrich_address(adr: dict, tries: int = 5, base_sleep: float = 1.2) -> dict:
    cur = dict(adr or {})
    for i in range(tries):
        try:
            cur = uzupelnij_braki_z_nominatim(cur)
            cur = dopelnij_powiat_gmina_jesli_brak(cur)
            return cur
        except Exception as e:
            if _is_remote_closed(e):
                time.sleep(base_sleep * (i + 1))
                continue
            raise
    print("[WARN] Nominatim: zrezygnowano po wielokrotnych próbach")
    return cur

PRICE_SELECTORS = [
    "strong[data-testid='ad-price']",
    "[data-testid='ad-price']",
    "[data-cy='ad-price']",
]
def _extract_price_from_dom(soup, wynik: dict):
    for sel in PRICE_SELECTORS:
        el = soup.select_one(sel)
        if el:
            txt = _norm_txt(el.get_text(" ", strip=True))
            if txt:
                wynik["cena"] = txt
                return

def _address_from_jsonld(soup) -> str:
    for s in soup.find_all("script", attrs={"type": "application/ld+json"}):
        raw = (s.string or "").strip()
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except Exception:
            try:
                data = json.loads(raw.replace("\n", " ").replace("\t", " "))
            except Exception:
                continue
        items = data if isinstance(data, list) else [data]
        for it in items:
            addr = it.get("address")
            if isinstance(addr, dict):
                parts = []
                street = addr.get("streetAddress")
                locality = addr.get("addressLocality") or addr.get("addressRegion")
                country = addr.get("addressCountry")
                if street:
                    parts.append(_norm_txt(street))
                if locality:
                    parts.append(_norm_txt(locality))
                if country:
                    parts.append(country if isinstance(country, str) else _norm_txt(country.get("name", "")))
                s = ", ".join([p for p in parts if p])
                if s:
                    return s
    return ""

def _pobierz_soup(url: str):
    opts = Options()
    opts.add_argument("--start-maximized")
    # opts.add_argument("--headless=new")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )

    driver = webdriver.Chrome(options=opts)
    try:
        try:
            driver.execute_cdp_cmd(
                "Page.addScriptToEvaluateOnNewDocument",
                {"source": "Object.defineProperty(navigator,'webdriver',{get:() => undefined})"},
            )
        except Exception:
            pass

        last_err = None
        for attempt in range(3):
            try:
                driver.get(url)
                break
            except Exception as e:
                last_err = e
                time.sleep(1.5 * (attempt + 1))
        if last_err and attempt == 2:
            raise last_err

        wait = WebDriverWait(driver, WAIT_SECONDS)
        try:
            wait.until(
                EC.any_of(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "meta[name='description']")),
                    EC.presence_of_element_located((By.CSS_SELECTOR, "strong[data-testid='ad-price']")),
                    EC.presence_of_element_located((By.CSS_SELECTOR, "link[rel='canonical']")),
                )
            )
        except Exception:
            time.sleep(min(5, WAIT_SECONDS))

        html = driver.page_source
    finally:
        driver.quit()
    return BeautifulSoup(html, "html.parser")

def _wyciagnij_adres_string(soup) -> str:
    mapa_link = soup.select_one("a[href='#map']") or soup.select_one("a[href*='mapa']") \
        or soup.find("span", attrs={"data-cy": "adPageAdLocalisation"})
    if mapa_link:
        t = mapa_link.get_text(strip=True)
        print("📍 Adres z UI:", t)
        return t
    return ""

def _fill_from_jsonld(obj, wynik: dict):
    if not isinstance(obj, (dict, list)):
        return
    items = obj if isinstance(obj, list) else [obj]
    for it in items:
        if not isinstance(it, dict):
            continue
        offers = it.get("offers") or it.get("offer")
        if isinstance(offers, dict):
            if "cena" not in wynik and offers.get("price"):
                wynik["cena"] = format_pln_int(offers.get("price")) or wynik.get("cena")
        floor_size = it.get("floorSize") or it.get("floorArea")
        if isinstance(floor_size, dict) and "metry" not in wynik:
            v = floor_size.get("value") or it.get("valueReference")
            m2 = fmt_metry(v)
            if m2:
                wynik["metry"] = m2
        if "liczba_pokoi" not in wynik and it.get("numberOfRooms"):
            wynik["liczba_pokoi"] = str(_num(str(it.get("numberOfRooms"))))
        if "rok_budowy" not in wynik and it.get("yearBuilt"):
            wynik["rok_budowy"] = str(it["yearBuilt"])
        if "material" not in wynik and it.get("material"):
            mat = it.get("material")
            if isinstance(mat, str):
                wynik["material"] = mat
        if "rynek" not in wynik and it.get("itemCondition"):
            cond = str(it["itemCondition"]).lower()
            if "new" in cond:
                wynik["rynek"] = "pierwotny"
            elif "used" in cond:
                wynik["rynek"] = "wtórny"

def pobierz_dane_z_otodom(url: str) -> dict:
    soup = _pobierz_soup(url)
    wynik = {}
    _extract_price_from_dom(soup, wynik)

    opis = soup.find("meta", attrs={"name": "description"})
    if opis and opis.has_attr("content"):
        content = opis["content"]
        if "cena" not in wynik:
            cena_match = re.search(r"za cenę ([\d\s]+zł)", content)
            if cena_match:
                wynik["cena"] = cena_match.group(1).strip()
        metry_match = re.search(r"ma ([\d.,]+ m²)", content)
        if metry_match:
            wynik["metry"] = metry_match.group(1).strip()
        pietro_match = re.search(r"na ([^,\.]+? piętrze)", content)
        if pietro_match:
            pietro_raw = pietro_match.group(1).strip().lower()
            if "parter" in pietro_raw:
                wynik["pietro"] = "parter"
            else:
                tylko_numer = re.search(r"\d+", pietro_raw)
                wynik["pietro"] = tylko_numer.group() if tylko_numer else ""

    for sel in [
        "strong[data-testid='ad-price-per-m']",
        "div[data-testid='ad-price-per-m']",
        "p[data-testid='ad-price-per-m']",
        "span[data-testid='ad-price-per-m']",
        "[data-testid='ad-price-per-m2']",
        "[data-cy='ad-price-per-m']",
    ]:
        el = soup.select_one(sel)
        if el:
            wynik["cena_za_metr"] = el.get_text(strip=True)
            break

    szczegoly = soup.select("div[data-testid='table-value']")
    for item in szczegoly:
        label_elem = item.select_one("div[data-testid='table-value-title']")
        value_elem = item.select_one("div[data-testid='table-value-value']")
        if label_elem and value_elem:
            label = label_elem.get_text(strip=True)
            value = value_elem.get_text(strip=True)
            if label in szukane_pola and szukane_pola[label] not in wynik:
                wynik[szukane_pola[label]] = value

    _extract_by_labels_generic(soup, wynik)

    try:
        for s in soup.find_all("script", attrs={"type": "application/ld+json"}):
            raw = (s.string or "").strip()
            if not raw:
                continue
            try:
                data = json.loads(raw)
                _fill_from_jsonld(data, wynik)
            except Exception:
                try:
                    data = json.loads(raw.replace("\n", " ").replace("\t", " "))
                    _fill_from_jsonld(data, wynik)
                except Exception:
                    pass
    except Exception:
        pass

    adres_string = _wyciagnij_adres_string(soup)
    adr_jsonld_str = _address_from_jsonld(soup)
    if adr_jsonld_str and len(adr_jsonld_str) > len(adres_string):
        adres_string = adr_jsonld_str

    adr = parsuj_adres_string(adres_string)
    try:
        adr = _safe_enrich_address(adr)
    except Exception as e:
        print(f"[WARN] Enrichment failed: {e}")

    if not (adr.get("ulica_nazwa") or "").strip():
        try:
            for s in soup.find_all("script", attrs={"type": "application/ld+json"}):
                raw = (s.string or "").strip()
                if not raw:
                    continue
                data = json.loads(raw.replace("\n", " ").replace("\t", " "))
                items = data if isinstance(data, list) else [data]
                found = False
                for it in items:
                    addr = it.get("address")
                    if isinstance(addr, dict):
                        street = addr.get("streetAddress")
                        if street:
                            nazwa = re.sub(r"\s+\d.*$", "", street).strip()
                            adr["ulica_nazwa"] = nazwa or street.strip()
                            found = True
                            break
                if found:
                    break
        except Exception:
            pass

    adr["gmina"] = _clean_gmina(adr.get("gmina"))
    dzielnica_csv = adr.get("dzielnica") or ""
    ulica_csv = adr.get("ulica_nazwa") or ""

    canonical = soup.find("link", rel="canonical")
    wynik["link"] = canonical["href"] if canonical and canonical.has_attr("href") else url

    if "cena_za_metr" not in wynik or not wynik["cena_za_metr"]:
        wynik["cena_za_metr"] = oblicz_cena_za_metr(wynik.get("cena", ""), wynik.get("metry", ""))

    miejscowosc = adr.get("miasto") or ""

    out = {
        "cena": wynik.get("cena", ""),
        "cena_za_metr": wynik.get("cena_za_metr", ""),
        "metry": wynik.get("metry", ""),
        "liczba_pokoi": wynik.get("liczba_pokoi", ""),
        "pietro": wynik.get("pietro", ""),
        "rynek": wynik.get("rynek", ""),
        "rok_budowy": wynik.get("rok_budowy", ""),
        "material": wynik.get("material", ""),
        "wojewodztwo": adr.get("wojewodztwo") or "",
        "powiat": adr.get("powiat") or "",
        "gmina": adr.get("gmina") or "",
        "miejscowosc": miejscowosc,
        "dzielnica": dzielnica_csv,
        "ulica": ulica_csv,
        "link": wynik["link"],
    }

    _consistency_pass_row(out)
    return out

# ============== ZAPIS DO JEDNEGO EXCELA (osobne arkusze) ==============
def _ensure_sheet_with_header(wb, sheetname: str, headers: list[str]):
    if sheetname in wb.sheetnames:
        ws = wb[sheetname]
        if ws.max_row < 1 or all((c.value in (None, "")) for c in ws[1]):
            for i, h in enumerate(headers, start=1):
                ws.cell(row=1, column=i, value=h)
        return ws
    ws = wb.create_sheet(title=sheetname)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    return ws

def write_rows_to_single_xlsx(sheetname: str, rows: list[dict]):
    if SINGLE_XLSX.exists():
        wb = load_workbook(SINGLE_XLSX)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])

    ws = _ensure_sheet_with_header(wb, sheetname, KOLUMNY_WYJ)

    for row in rows:
        ws.append([row.get(col, "") for col in KOLUMNY_WYJ])

    SINGLE_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SINGLE_XLSX)

# ============== PIPE: czytaj linki (CSV) -> scrap -> JEDEN XLSX ==============
def przetworz_linki_z_intake_csv_do_jednego_xlsx(input_csv: Path, sheetname: str):
    dane_lista = []
    with input_csv.open(mode="r", encoding="utf-8") as plik_wejsciowy:
        reader = csv.reader(plik_wejsciowy)
        next(reader, None)  # nagłówek
        for row in reader:
            if not row or not row[0].strip():
                continue
            link = row[0].strip()
            print(f"\n🌐 Przetwarzam link: {link}")
            dane = None
            for attempt in range(2):
                try:
                    dane = pobierz_dane_z_otodom(link)
                    break
                except Exception as e:
                    if _is_remote_closed(e) and attempt == 0:
                        print(f"[WARN] Tymczasowy problem sieciowy: {e} — ponawiam…")
                        time.sleep(2.0)
                        continue
                    print(f"❌ Błąd przy linku {link}: {e}")
                    dane = None
                    break
            if not dane:
                continue
            for k in KOLUMNY_WYJ:
                dane.setdefault(k, "")
            dane_lista.append(dane)

    dane_lista = [r for r in dane_lista if (r.get("cena") or "").strip()]

    if not dane_lista:
        # upewnij się, że plik i arkusz istnieją z nagłówkiem
        write_rows_to_single_xlsx(sheetname, [])
        print(f"\n⚠️ Brak rekordów z ceną — utworzono/utrzymano arkusz z nagłówkiem w {SINGLE_XLSX.resolve()}.")
    else:
        write_rows_to_single_xlsx(sheetname, dane_lista)
        print(f"\n✅ Zapisano {len}(dane_lista) ogłoszeń do {SINGLE_XLSX.resolve()} (arkusz: {sheetname}).")

# ------------------------------ CLI ------------------------------
if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--region", "-r", required=True, help="Slug województwa, np. 'mazowieckie'")
    args = ap.parse_args()

    slug = args.region.strip().lower()
    input_csv = BASE_LINKI / f"intake_{slug}.csv"   # czytamy z Desktop/Pulpit/baza danych/linki
    sheetname  = slug                                # zapisujemy do jednego pliku, arkusz = slug

    if not input_csv.exists():
        raise SystemExit(f"Nie znaleziono pliku linków: {input_csv}")

    przetworz_linki_z_intake_csv_do_jednego_xlsx(input_csv, sheetname)
