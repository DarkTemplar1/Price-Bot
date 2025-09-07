# -*- coding: utf-8 -*-
import sys
import csv
from pathlib import Path
from typing import List
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook

RAPORT_SHEET = "raport"
RAPORT_ODF = "raport_odfiltrowane"

REQ_COLS: List[str] = [
    "Nr KW","Typ Księgi","Stan Księgi","Województwo","Powiat","Gmina",
    "Miejscowość","Dzielnica","Położenie","Nr działek po średniku","Obręb po średniku",
    "Ulica","Sposób korzystania","Obszar","Ulica(dla budynku)",
    "przeznaczenie (dla budynku)","Ulica(dla lokalu)","Nr budynku( dla lokalu)",
    "Przeznaczenie (dla lokalu)","Cały adres (dla lokalu)","Czy udziały?"
]

VALUE_COLS: List[str] = [
    "Średnia cena za m2 ( z bazy)",
    "Średnia skorygowana cena za m2",
    "Statyczna wartość nieruchomości"
]

# Nagłówek plików wynikowych CSV (tak jak w scraperach)
WYNIKI_HEADER: List[str] = [
    "cena","cena_za_metr","metry","liczba_pokoi","pietro","rynek","rok_budowy","material",
    "wojewodztwo","powiat","gmina","miejscowosc","dzielnica","ulica","link",
]

SUPPORTED = {".xlsx", ".xlsm"}

# --- REGIONY: pracujemy na etykietach (PL) i slugach (ASCII) ---
VOIVODESHIPS_LABEL_SLUG: list[tuple[str, str]] = [
    ("Dolnośląskie", "dolnoslaskie"),
    ("Kujawsko-Pomorskie", "kujawsko-pomorskie"),
    ("Lubelskie", "lubelskie"),
    ("Lubuskie", "lubuskie"),
    ("Łódzkie", "lodzkie"),
    ("Małopolskie", "malopolskie"),
    ("Mazowieckie", "mazowieckie"),
    ("Opolskie", "opolskie"),
    ("Podkarpackie", "podkarpackie"),
    ("Podlaskie", "podlaskie"),
    ("Pomorskie", "pomorskie"),
    ("Śląskie", "slaskie"),
    ("Świętokrzyskie", "swietokrzyskie"),
    ("Warmińsko-Mazurskie", "warminsko-mazurskie"),
    ("Wielkopolskie", "wielkopolskie"),
    ("Zachodniopomorskie", "zachodniopomorskie"),
]

# --------------------- Desktop/Pulpit ---------------------
def _detect_desktop() -> Path:
    home = Path.home()
    for name in ("Desktop", "Pulpit"):
        p = home / name
        if p.exists():
            return p
    return home

def ensure_base_dirs() -> Path:
    desktop = _detect_desktop()
    base = desktop / "baza danych"
    (base / "linki").mkdir(parents=True, exist_ok=True)
    (base / "województwa").mkdir(parents=True, exist_ok=True)
    return base
# ---------------------------------------------------------

# --------------------- CSV helpery -----------------------
def _ensure_csv(path: Path, header: List[str]) -> bool:
    if path.exists():
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        if header:
            w.writerow(header)
    return True

def create_voivodeship_csvs(base: Path) -> dict:
    created = {"linki": 0, "województwa": 0}
    linki_dir = base / "linki"
    woj_dir = base / "województwa"
    for label, _slug in VOIVODESHIPS_LABEL_SLUG:
        if _ensure_csv(linki_dir / f"{label}.csv", ["link"]):
            created["linki"] += 1
        if _ensure_csv(woj_dir / f"{label}.csv", WYNIKI_HEADER):
            created["województwa"] += 1
    return created
# ---------------------------------------------------------

def _ensure_base_headers(ws) -> None:
    if ws.max_row == 1 and ws.max_column == 1 and (ws.cell(1,1).value in (None, "")):
        for c, h in enumerate(REQ_COLS, start=1):
            ws.cell(row=1, column=c, value=h)
        return
    existing = [cell.value or "" for cell in ws[1]]
    for col in REQ_COLS:
        if col not in existing:
            ws.cell(row=1, column=ws.max_column + 1, value=col)

def _ensure_value_cols_create_if_missing(ws, anchor: str = "Czy udziały?") -> None:
    """
    Sprawdza, czy kolumny VALUE_COLS już istnieją (gdziekolwiek w wierszu nagłówków).
    - Jeśli istnieją, nic nie dodaje (brak duplikatów).
    - Jeśli którejś brakuje, DODAJE brakujące kolumny tuż za kolumną 'anchor'
      i ustawia ich nagłówki w kolejności z VALUE_COLS.
    """
    headers = [cell.value or "" for cell in ws[1]]

    if anchor in headers:
        anchor_idx = headers.index(anchor) + 1  # 1-based
    else:
        ws.cell(row=1, column=ws.max_column + 1, value=anchor)
        anchor_idx = ws.max_column  # nowy anchor dodany na końcu

    # odśwież listę nagłówków po ewentualnym dodaniu anchor
    headers = [cell.value or "" for cell in ws[1]]
    existing = set(headers)

    missing = [name for name in VALUE_COLS if name not in existing]
    if not missing:
        return

    ws.insert_cols(anchor_idx + 1, amount=len(missing))
    for i, name in enumerate(missing, start=1):
        ws.cell(row=1, column=anchor_idx + i, value=name)

def ensure_sheet_and_columns(xlsx_path: Path) -> None:
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()
        wb.active.title = RAPORT_SHEET
    for sheet_name in (RAPORT_SHEET, RAPORT_ODF):
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        _ensure_base_headers(ws)
        _ensure_value_cols_create_if_missing(ws, anchor="Czy udziały?")
    wb.save(xlsx_path)

def pick_file_via_gui() -> Path:
    root = tk.Tk(); root.withdraw()
    p = filedialog.askopenfilename(title="Wybierz plik Excel",
                                   filetypes=[("Excel", "*.xlsx *.xlsm"), ("Wszystkie pliki","*.*")])
    root.destroy()
    if not p:
        raise SystemExit(0)
    return Path(p)

def info(msg: str) -> None:
    try:
        r = tk.Tk(); r.withdraw()
        messagebox.showinfo("Gotowe", msg, parent=r)
        r.destroy()
    except Exception:
        print(msg)

def error(msg: str) -> None:
    try:
        r = tk.Tk(); r.withdraw()
        messagebox.showerror("Błąd", msg, parent=r)
        r.destroy()
    except Exception:
        print("BŁĄD:", msg, file=sys.stderr)

def main():
    try:
        base_dir = ensure_base_dirs()
    except Exception as e:
        error(f"Nie udało się utworzyć folderów na Pulpicie/Desktop:\n{e}")
        raise SystemExit(3)

    try:
        created = create_voivodeship_csvs(base_dir)
    except Exception as e:
        error(f"Nie udało się utworzyć plików CSV dla województw:\n{e}")
        raise SystemExit(4)

    if len(sys.argv) >= 2:
        xlsx_path = Path(sys.argv[1])
    else:
        xlsx_path = pick_file_via_gui()

    if xlsx_path.suffix.lower() not in SUPPORTED:
        error(f"Nieobsługiwane rozszerzenie: {xlsx_path.suffix}. Zapisz plik jako .xlsx lub .xlsm.")
        raise SystemExit(2)

    try:
        ensure_sheet_and_columns(xlsx_path)
        info(
            "Przygotowano arkusze 'raport' i 'raport_odfiltrowane' oraz dodano brakujące kolumny za 'Czy udziały?'\n"
            f"Plik: {xlsx_path}\n\n"
            "Utworzono/zweryfikowano również strukturę folderów i CSV per województwo (etykiety):\n"
            f"• {base_dir/'linki'} — nowych plików: {created.get('linki',0)}\n"
            f"• {base_dir/'województwa'} — nowych plików: {created.get('województwa',0)}"
        )
    except Exception as e:
        error(f"Nie udało się przygotować pliku:\n{e}")
        raise SystemExit(1)

if __name__ == "__main__":
    main()
