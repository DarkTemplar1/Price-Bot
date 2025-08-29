# -*- coding: utf-8 -*-
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from openpyxl import load_workbook

APP_TITLE = "testy0703 — Uzupełnianie z TERYT.xlsx"
TERYT_DEFAULT = "TERYT.xlsx"

# --- normalizacja ---
DIAC_MAP = str.maketrans({
    "ą":"a","ć":"c","ę":"e","ł":"l","ń":"n","ó":"o","ś":"s","ź":"z","ż":"z",
    "Ą":"A","Ć":"C","Ę":"E","Ł":"L","Ń":"N","Ó":"O","Ś":"S","Ź":"Z","Ż":"Z",
})
def norm_key(s: str) -> str:
    if s is None: return ""
    s = str(s).strip().lower()
    s = re.sub(r"^(wojew[oó]dztwo|woj\.)\s+", "", s)  # usuń "województwo " / "woj. "
    return s.translate(DIAC_MAP)

# --- stara->nowa nazwa województwa (klucze znormalizowane) ---
OLD_TO_NEW_VOIV_RAW = {
    "białostockie": "podlaskie","bielskie": "śląskie","bydgoskie": "kujawsko-pomorskie",
    "chełmskie": "lubelskie","ciechanowskie": "mazowieckie","częstochowskie": "śląskie",
    "elbląskie": "warmińsko-mazurskie","gdańskie": "pomorskie","gorzowskie": "lubuskie",
    "jeleniogórskie": "dolnośląskie","kaliskie": "wielkopolskie","katowickie": "śląskie",
    "kieleckie": "świętokrzyskie","koszalińskie": "zachodniopomorskie","krakowskie": "małopolskie",
    "krośnieńskie": "podkarpackie","legnickie": "dolnośląskie","leszczyńskie": "wielkopolskie",
    "łomżyńskie": "podlaskie","nowosądeckie": "małopolskie","olsztyńskie": "warmińsko-mazurskie",
    "opole": "opolskie","ostrołęckie": "mazowieckie","piotrkowskie": "łódzkie",
    "płockie": "mazowieckie","poznańskie": "wielkopolskie","przemyskie": "podkarpackie",
    "radomskie": "mazowieckie","rzeszowskie": "podkarpackie","siedleckie": "mazowieckie",
    "sieradzkie": "łódzkie","skierniewickie": "łódzkie","słupskie": "pomorskie",
    "suwałskie": "podlaskie","szczecińskie": "zachodniopomorskie","tarnobrzeskie": "podkarpackie",
    "tarnowskie": "małopolskie","toruńskie": "kujawsko-pomorskie","wałbrzyskie": "dolnośląskie",
    "warszawskie": "mazowieckie","włocławskie": "kujawsko-pomorskie","wrocławskie": "dolnośląskie",
    "zamojsko-chełmskie": "lubelskie","zielonogórskie": "lubuskie"
}
OLD_TO_NEW_VOIV = {norm_key(k): v for k, v in OLD_TO_NEW_VOIV_RAW.items()}

def map_old_voiv(name: str) -> str:
    n = norm_key(name)
    return OLD_TO_NEW_VOIV.get(n, name)

# --- TERYT wczytanie ---
def load_teryt(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["TERYT"] if "TERYT" in wb.sheetnames else wb.active
    headers = {str(c.value).strip(): i for i, c in enumerate(ws[1], start=1) if c.value}
    req = ["Województwo","Powiat","Gmina","Miejscowość","Dzielnica"]
    for r in req:
        if r not in headers:
            raise RuntimeError(f"Brak kolumny '{r}' w {path}")
    rows = []
    for r in range(2, ws.max_row+1):
        rows.append({k: (ws.cell(row=r, column=headers[k]).value or "") for k in req})
    return rows

# --- dopasowania ---
def filter_candidates(teryt_rows, w=None, p=None, g=None, m=None, d=None):
    w = norm_key(map_old_voiv(w)) if w else ""
    p = norm_key(p) if p else ""
    g = norm_key(g) if g else ""
    m = norm_key(m) if m else ""
    d = norm_key(d) if d else ""
    out = []
    for rec in teryt_rows:
        ok = True
        if w and norm_key(rec["Województwo"]) != w: ok = False
        if p and norm_key(rec["Powiat"]) != p: ok = False
        if g and norm_key(rec["Gmina"]) != g: ok = False
        if m and norm_key(rec["Miejscowość"]) != m: ok = False
        if d and norm_key(rec["Dzielnica"]) != d: ok = False
        if ok: out.append(rec)
    return out

def stable_values(cands):
    fields = ["Województwo","Powiat","Gmina","Miejscowość","Dzielnica"]
    out = {}
    for f in fields:
        vals = {rec[f] for rec in cands if str(rec[f]).strip() != ""}
        if len(vals) == 1:
            out[f] = next(iter(vals))
    return out

# --- GUI ---
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("980x600")
        self.teryt_path = tk.StringVar(value=TERYT_DEFAULT if Path(TERYT_DEFAULT).exists() else "")
        top = tk.Frame(self); top.pack(fill="x", padx=10, pady=6)
        tk.Label(top, text="TERYT.xlsx:").pack(side="left")
        tk.Entry(top, textvariable=self.teryt_path, width=60).pack(side="left", padx=6)
        tk.Button(top, text="Wybierz…", command=self.pick_teryt).pack(side="left")
        self.n_rows_var = tk.StringVar(value="5")
        tk.Label(top, text="Ilość wierszy:").pack(side="left", padx=(20,4))
        tk.Entry(top, textvariable=self.n_rows_var, width=5).pack(side="left")
        tk.Button(top, text="Generuj pola", command=self.gen_rows).pack(side="left", padx=8)
        tk.Button(top, text="Sprawdzam (uzupełnij z TERYT)", command=self.check_and_fill).pack(side="right")
        self.canvas = tk.Canvas(self, borderwidth=0)
        self.frame = tk.Frame(self.canvas)
        self.scroll = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scroll.set)
        self.scroll.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas_frame = self.canvas.create_window((0,0), window=self.frame, anchor="nw")
        self.frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        headers = ["Województwo","Powiat","Gmina","Miejscowość","Dzielnica"]
        self.entries = []
        head = tk.Frame(self.frame); head.pack(fill="x")
        for i, h in enumerate(headers):
            tk.Label(head, text=h, font=("TkDefaultFont", 10, "bold"), width=22, anchor="w").grid(row=0, column=i, padx=4, pady=4, sticky="w")
        self.rows_container = tk.Frame(self.frame); self.rows_container.pack(fill="both", expand=True)
        self.gen_rows()

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_frame, width=event.width)

    def pick_teryt(self):
        path = filedialog.askopenfilename(title="Wskaż TERYT.xlsx", filetypes=[("Excel","*.xlsx")])
        if path:
            self.teryt_path.set(path)

    def _clear_rows(self):
        for w in self.rows_container.winfo_children():
            w.destroy()
        self.entries.clear()

    def gen_rows(self):
        try:
            n = int(self.n_rows_var.get())
            n = max(1, min(500, n))
        except Exception:
            messagebox.showwarning("Uwaga", "Podaj poprawną liczbę wierszy (1–500).")
            return
        self._clear_rows()
        for _ in range(n):
            rowf = tk.Frame(self.rows_container)
            rowf.pack(fill="x", padx=4, pady=2)
            row_entries = []
            for _c in range(5):
                e = tk.Entry(rowf, width=24)
                e.grid(row=0, column=_c, padx=4, pady=2, sticky="w")
                row_entries.append(e)
            self.entries.append(row_entries)

    def check_and_fill(self):
        path = self.teryt_path.get()
        if not path or not Path(path).exists():
            messagebox.showerror("Błąd", "Wskaż plik TERYT.xlsx.")
            return
        try:
            trows = load_teryt(path)
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie udało się wczytać TERYT.xlsx: {e}")
            return

        updated = 0
        for row_entries in self.entries:
            w, p, g, m, d = [e.get().strip() for e in row_entries]

            # 0) od razu zamień starą nazwę woj. w polu (jeśli trzeba)
            if w:
                mapped = map_old_voiv(w)
                if mapped != w:
                    row_entries[0].delete(0, "end"); row_entries[0].insert(0, mapped)
                    w = mapped  # aktualny tekst w polu

            # A: pełne filtry
            cands = filter_candidates(trows, w, p, g, m, d)
            # B: woj + miejscowość
            if not cands and (w or m):
                cands = filter_candidates(trows, w, None, None, m, None)
            # C: sama miejscowość
            if not cands and m:
                cands = filter_candidates(trows, None, None, None, m, None)
            # D: sama gmina
            if not cands and g:
                cands = filter_candidates(trows, None, None, g, None, None)

            if len(cands) == 1:
                # === ZAMIANA: nadpisz WSZYSTKIE pola kanonicznymi wartościami ===
                rec = cands[0]
                vals = [rec["Województwo"], rec["Powiat"], rec["Gmina"], rec["Miejscowość"], rec["Dzielnica"]]
                for e, v in zip(row_entries, vals):
                    e.delete(0, "end"); e.insert(0, str(v))
                updated += 1
            elif len(cands) > 1:
                # === ZAMIANA TYLKO JEDNOZNACZNYCH PÓL (wspólne dla wszystkich kandydatów) ===
                stab = stable_values(cands)
                changed = False
                for (e, key) in zip(row_entries, ["Województwo","Powiat","Gmina","Miejscowość","Dzielnica"]):
                    if key in stab:
                        e.delete(0, "end"); e.insert(0, str(stab[key])); changed = True
                if changed:
                    updated += 1

        messagebox.showinfo("Zakończono", f"Uzupełniono {updated} wierszy (w całości lub częściowo).")

if __name__ == "__main__":
    App().mainloop()
