# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd
from openpyxl import load_workbook, Workbook

# ==========================
# Stałe dostępne dla UI / filtrów
# ==========================
KOL_UDZIALY = "Czy udziały ?"
KOL_LOKAL = "Przeznaczenie (dla lokalu)"
KOL_KW = "Nr KW"
SHEET_ROBOCZY = "roboczy"

SHEET_RAPORT = "raport"
SHEET_RAPORT_ODF = "raport_odfiltrowane"

# kolumny adresowe wymagane w „raport”
ADDRESS_COLUMNS_ORDER: list[str] = [
    "Nr KW",
    "Typ Księgi",
    "Stan Księgi",
    "Województwo",
    "Powiat",
    "Gmina",
    "Miejscowość",
    "Dzielnica",
    "Położenie",
    "Nr działek po średniku",
    "Obręb po średniku",
    "Ulica",
    "Sposób korzystania",
    "Obszar",
    "Ulica(dla budynku)",
    "Przeznaczenie (dla budynku)",
    "Ulica(dla lokalu)",
    "Nr budynku( dla lokalu)",
    "Przeznaczenie (dla lokalu)",
    "Cały adres (dla lokalu)",
    "Czy udziały ?",
]

# kolumny dla bazy „Baza_danych.xlsx” / arkusza „Mieszkania”
BAZA_MIESZKANIA_HEADERS: list[str] = [
    "cena",
    "cena_za_metr",
    "metry",
    "liczba_pokoi",
    "pietro",
    "rynek",
    "rok_budowy",
    "material",
    "wojewodztwo",
    "powiat",
    "gmina",
    "miejscowosc",
    "dzielnica",
    "ulica",
    "link",
]


# ==========================
# Proste utilsy Excela (roboczy)
# ==========================
def excel_first_sheet_name(file_path: Path) -> str:
    """Zwróć nazwę pierwszego arkusza w skoroszycie."""
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    if not xls.sheet_names:
        raise ValueError("Plik nie zawiera żadnych arkuszy.")
    return xls.sheet_names[0]


def clone_first_sheet_to_roboczy(file_path: Path, source_sheet: str) -> None:
    """
    Usuń istniejący arkusz 'roboczy' i utwórz go na nowo
    jako pełną kopię arkusza źródłowego (z formatowaniem i formułami).
    """
    wb = load_workbook(file_path)
    if source_sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Brak arkusza źródłowego: {source_sheet}")

    if SHEET_ROBOCZY in wb.sheetnames:
        wb.remove(wb[SHEET_ROBOCZY])

    ws_src = wb[source_sheet]
    ws_new = wb.copy_worksheet(ws_src)
    ws_new.title = SHEET_ROBOCZY

    wb.save(file_path)
    wb.close()


def ensure_roboczy_on_start(file_path: Path, source_sheet: str) -> None:
    wb = load_workbook(file_path)
    has_roboczy = SHEET_ROBOCZY in wb.sheetnames
    wb.close()
    if not has_roboczy:
        clone_first_sheet_to_roboczy(file_path, source_sheet)


def _coerce_text_cols(df: pd.DataFrame) -> pd.DataFrame:
    for col in (KOL_UDZIALY, KOL_LOKAL, KOL_KW):
        if col in df.columns:
            df[col] = df[col].astype("string")
    return df


def read_original_and_view(file_path: Path, source_sheet: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    df_original = pd.read_excel(file_path, sheet_name=source_sheet, engine="openpyxl")
    df_original = _coerce_text_cols(df_original)

    try:
        df_view = pd.read_excel(file_path, sheet_name=SHEET_ROBOCZY, engine="openpyxl")
        df_view = _coerce_text_cols(df_view)
    except ValueError:
        # brak arkusza 'roboczy'
        df_view = df_original.copy()

    return df_original, df_view


def write_view_to_roboczy(file_path: Path, df_view: pd.DataFrame) -> None:
    """
    Zapisz df_view do arkusza 'roboczy' (wartości; formatowanie nie jest zachowywane).
    """
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_view.to_excel(writer, sheet_name=SHEET_ROBOCZY, index=False)


# ==========================
# Walidacja/utworzenie plików/arkuszy z nagłówkami
# ==========================
def ensure_workbook_and_sheet_with_header(
    file_path: Path,
    sheet_name: str,
    headers: Sequence[str],
) -> None:
    """
    Zapewnij istnienie pliku Excela, arkusza i wiersza nagłówków.
    Jeśli arkusz istnieje, dołóż brakujące kolumny na końcu.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(list(headers))
        wb.save(file_path)
        wb.close()
        return

    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(list(headers))
        wb.save(file_path)
        wb.close()
        return

    # arkusz istnieje – uzupełnij brakujące kolumny
    ws = wb[sheet_name]
    existing = [c.value if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    existing = [str(x) for x in existing]
    if all(v == "" for v in existing):
        # pusty nagłówek – nadpisz
        ws.delete_rows(1, 1)
        ws.append(list(headers))
        wb.save(file_path)
        wb.close()
        return

    # dołóż brakujące
    missing = [h for h in headers if h not in existing]
    if missing:
        # dopisz puste komórki nagłówka na końcu
        for i, name in enumerate(missing, start=len(existing) + 1):
            ws.cell(row=1, column=i, value=name)
    wb.save(file_path)
    wb.close()


def append_rows_dicts(
    file_path: Path,
    sheet_name: str,
    rows: Iterable[dict],
    headers: Sequence[str],
) -> None:
    """
    Dopisz wiersze (dict) do istniejącego arkusza – zgodnie z kolejnością `headers`.
    Brakujące klucze traktowane są jako puste.
    """
    ensure_workbook_and_sheet_with_header(file_path, sheet_name, headers)

    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    wb.save(file_path)
    wb.close()


# ==========================
# Narzędzia do arkuszy: RAPORT / RAPORT_ODFILTROWANE
# ==========================
def _norm_text(s: str | None) -> str:
    import re, unicodedata

    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.casefold()
    s = re.sub(r"[ \t\r\n\.\-\_\?\(\)]", "", s)
    return s


def ensure_raport_odfiltrowane(file_path: Path, raport_sheet: str = SHEET_RAPORT) -> None:
    """
    Utwórz arkusz 'raport_odfiltrowane' jeśli nie istnieje i skopiuj wiersz nagłówka
    z arkusza 'raport'.
    """
    file_path = Path(file_path)
    wb = load_workbook(file_path)
    if raport_sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Brak arkusza źródłowego: {raport_sheet}")

    ws_src = wb[raport_sheet]
    headers = [c.value for c in next(ws_src.iter_rows(min_row=1, max_row=1))]
    headers = [h if h is not None else "" for h in headers]

    if SHEET_RAPORT_ODF not in wb.sheetnames:
        ws_new = wb.create_sheet(SHEET_RAPORT_ODF)
        ws_new.append(headers)
        wb.save(file_path)
        wb.close()
        return

    # jeśli istnieje – upewnij nagłówki (dołóż brakujące)
    ws = wb[SHEET_RAPORT_ODF]
    cur = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    cur = [x if x is not None else "" for x in cur]
    missing = [h for h in headers if h not in cur]
    for i, name in enumerate(missing, start=len(cur) + 1):
        ws.cell(row=1, column=i, value=name)

    wb.save(file_path)
    wb.close()


def _read_sheet_df(file_path: Path, sheet: str) -> pd.DataFrame:
    return pd.read_excel(file_path, sheet_name=sheet, engine="openpyxl")


def _write_sheet_df(file_path: Path, sheet: str, df: pd.DataFrame) -> None:
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet, index=False)


def filter_raport_move_to_odf(file_path: Path, mode: str) -> dict:
    """
    Wykonaj operację filtrowania/przerzucania między 'raport' a 'raport_odfiltrowane'.

    mode:
      - 'brak'   → nic nie rób
      - 'jeden'  → w 'raport' znajdź wiersze gdzie KOL_UDZIALY == 'tak' (case-insens) i przerzuć do ODF
      - 'lokal'  → w 'raport' znajdź wiersze gdzie KOL_LOKAL != 'LOKAL MIESZKALNY' i przerzuć do ODF
      - 'oba'    → przenieś wiersze spełniające którykolwiek z ww. warunków
      - 'cofnij' → przenieś WSZYSTKIE wiersze z ODF z powrotem do 'raport' i wyczyść ODF (zostaw nagłówek)

    Zwraca słownik ze statystykami: {"moved_to_odf": int, "returned_to_raport": int}
    """
    file_path = Path(file_path)
    ensure_raport_odfiltrowane(file_path, SHEET_RAPORT)

    if mode == "brak":
        return {"moved_to_odf": 0, "returned_to_raport": 0}

    if mode == "cofnij":
        # przenieś wszystkie z ODF do raport
        df_r = _read_sheet_df(file_path, SHEET_RAPORT)
        try:
            df_odf = _read_sheet_df(file_path, SHEET_RAPORT_ODF)
        except ValueError:
            df_odf = pd.DataFrame(columns=list(df_r.columns))

        # usuń nagłówek logiczny – DataFrame już go nie ma, ale upewnij kolumny
        df_odf = df_odf[df_odf.notna().any(axis=1)]

        returned = len(df_odf)
        if returned:
            # dopasuj kolumny
            for c in df_r.columns:
                if c not in df_odf.columns:
                    df_odf[c] = pd.NA
            df_odf = df_odf[df_r.columns]
            df_r2 = pd.concat([df_r, df_odf], ignore_index=True)
        else:
            df_r2 = df_r

        # zapisz: raport = scalony, odf = tylko nagłówek
        _write_sheet_df(file_path, SHEET_RAPORT, df_r2)
        _write_sheet_df(file_path, SHEET_RAPORT_ODF, pd.DataFrame(columns=list(df_r2.columns)))
        return {"moved_to_odf": 0, "returned_to_raport": returned}

    # pozostałe tryby – przerzucamy z raport → odfiltrowane
    df_r = _read_sheet_df(file_path, SHEET_RAPORT)
    try:
        df_odf = _read_sheet_df(file_path, SHEET_RAPORT_ODF)
    except ValueError:
        df_odf = pd.DataFrame(columns=list(df_r.columns))

    def _col(df: pd.DataFrame, wanted: str) -> str | None:
        target = _norm_text(wanted)
        mapping = {c: _norm_text(str(c)) for c in df.columns}
        for c, n in mapping.items():
            if n == target or target in n:
                return c
        return None

    col_ud = _col(df_r, KOL_UDZIALY)
    col_lok = _col(df_r, KOL_LOKAL)

    mask_jeden = pd.Series([False] * len(df_r))
    if col_ud:
        ser = df_r[col_ud].astype("string").fillna("").str.strip().str.casefold()
        mask_jeden = ser == "tak"

    def _is_lokal_mieszkalny(x) -> bool:
        return _norm_text(str(x)) == "lokalmieszkalny"

    mask_lokal = pd.Series([False] * len(df_r))
    if col_lok:
        mask_lokal = ~df_r[col_lok].apply(_is_lokal_mieszkalny)

    if mode == "jeden":
        to_move = mask_jeden
    elif mode == "lokal":
        to_move = mask_lokal
    elif mode == "oba":
        to_move = mask_jeden | mask_lokal
    else:
        raise ValueError("Nieznany tryb filtrowania: " + str(mode))

    df_move = df_r[to_move].copy()
    moved = len(df_move)
    if moved == 0:
        return {"moved_to_odf": 0, "returned_to_raport": 0}

    # dopasuj kolumny i dołącz do ODF
    for c in df_r.columns:
        if c not in df_odf.columns:
            df_odf[c] = pd.NA
    df_move = df_move[df_odf.columns] if set(df_odf.columns) == set(df_r.columns) else df_move[df_r.columns]
    df_odf2 = pd.concat([df_odf, df_move], ignore_index=True)

    # usuń z raportu przenoszone (tu: wyrzucamy całe wiersze)
    df_keep = df_r[~to_move].copy()

    _write_sheet_df(file_path, SHEET_RAPORT, df_keep)
    _write_sheet_df(file_path, SHEET_RAPORT_ODF, df_odf2)

    return {"moved_to_odf": moved, "returned_to_raport": 0}


# ==========================
# Kolumna „Położenie” – składanie/uzupełnianie
# ==========================
def compose_polozenie(
    woj: str = "",
    powiat: str = "",
    gmina: str = "",
    miejscowosc: str = "",
    dzielnica: str = "",
    ulica: str = "",
) -> str:
    """Złożenie pola 'Położenie' jako 'Województwo; Powiat; Gmina; Miejscowość; Dzielnica; Ulica'."""
    parts = [woj, powiat, gmina, miejscowosc, dzielnica, ulica]
    cleaned = [str(p).strip() for p in parts if str(p).strip()]
    return "; ".join(cleaned)


def update_polozenie_column(file_path: Path, sheet: str = SHEET_RAPORT) -> int:
    """
    Uzupełnij/odśwież kolumnę 'Położenie' w podanym arkuszu na podstawie kolumn:
    Województwo, Powiat, Gmina, Miejscowość, Dzielnica, Ulica.

    Zwraca liczbę zaktualizowanych wierszy.
    """
    df = _read_sheet_df(file_path, sheet)
    changed = 0
    for col in ["Położenie", "Województwo", "Powiat", "Gmina", "Miejscowość", "Dzielnica", "Ulica"]:
        if col not in df.columns:
            # dołóż brakujące kolumny
            df[col] = ""

    new_vals = []
    for _, r in df.iterrows():
        s = compose_polozenie(r["Województwo"], r["Powiat"], r["Gmina"], r["Miejscowość"], r["Dzielnica"], r["Ulica"])
        new_vals.append(s)
    if "Położenie" in df.columns:
        changed = int((df["Położenie"].fillna("") != pd.Series(new_vals)).sum())
    df["Położenie"] = new_vals
    _write_sheet_df(file_path, sheet, df)
    return changed


def reorder_columns_by_header(
    file_path: Path,
    sheet: str,
    desired_order: Sequence[str],
    keep_others_at_end: bool = True,
) -> None:
    """
    Ułóż kolumny w arkuszu wg `desired_order`. Brakujące dopisz puste, nieznane zostaw na końcu (domyślnie).
    """
    df = _read_sheet_df(file_path, sheet)
    # dołóż brakujące
    for c in desired_order:
        if c not in df.columns:
            df[c] = pd.NA
    ordered = [c for c in desired_order if c in df.columns]
    if keep_others_at_end:
        rest = [c for c in df.columns if c not in ordered]
        new_cols = ordered + rest
    else:
        new_cols = ordered
    df = df[new_cols]
    _write_sheet_df(file_path, sheet, df)


# ==========================
# Baza danych: upewnij plik/arkusz/nagłówki i dopisz dane
# ==========================
def ensure_baza_mieszkania(file_path: Path, sheet: str = "Mieszkania") -> None:
    """
    Zapewnij plik Excela dla bazy danych i arkusz z wymaganymi kolumnami.
    """
    ensure_workbook_and_sheet_with_header(file_path, sheet, BAZA_MIESZKANIA_HEADERS)


def append_mieszkania_rows(file_path: Path, rows: Iterable[dict], sheet: str = "Mieszkania") -> None:
    """
    Dopisz rekordy ogłoszeń do arkusza 'Mieszkania' w 'Baza_danych.xlsx'.
    """
    append_rows_dicts(file_path, sheet, rows, BAZA_MIESZKANIA_HEADERS)
